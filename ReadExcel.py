import openpyxl

path = "E:\\edureka\\Assignment9\\testdata\\loginData.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook["Sheet2"]
rows = sheet.max_row
columns = sheet.max_column

print(rows)
print(columns)

for i in range(1, rows+1):
    for j in range(1, columns+1):
        print(sheet.cell(i, j).value, end=' ')

    print()
